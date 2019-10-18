# -*- coding: utf-8 -*-
import requests
from io import BytesIO
from pyzbar import pyzbar
from PIL import Image
import openpyxl
import os
import sys
import threadpool

# i = 0


def openpyxl_write(path, star_index, end_index, pool):
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.worksheets[0]
    # list(worksheet.rows)[0:10]
    # list(worksheet.rows)[star_index:end_index]
    data = []
    for index, row in enumerate(worksheet.rows):
        # print index
        if index >= star_index:
            if index > end_index:
                break
            else:
                data.append({'row0': row[2], 'row1': row[14]})
    requests = threadpool.makeRequests(test, data)
    [pool.putRequest(req) for req in requests]
    pool.wait()
    print "开始保存"
    workbook.save(filename=path)
    print "结束保存"


def test(row):
    # global i
    # print i
    # i = i + 1
    url = row["row0"].value
    if url != "二维码URL":
        try:
            data_url = get_ewm(url)
            row["row1"].value = data_url
        except BaseException as e:
            try:
                data_url = get_ewm(url)
                row["row1"].value = data_url
            except  BaseException as e:
                print  url


def get_ewm(img_adds):
    """ 读取二维码的内容： img_adds：二维码地址（可以是网址也可是本地地址 """
    if os.path.isfile(img_adds):
        # 从本地加载二维码图片
        img = Image.open(img_adds)
    else:
        # 从网络下载并加载二维码图片
        rq_img = requests.get(img_adds).content
        img = Image.open(BytesIO(rq_img))
        # img.show()  # 显示图片，测试用
    txt_list = pyzbar.decode(img)
    for txt in txt_list:
        barcodeData = txt.data.decode("utf-8")
        return barcodeData


def bootstarp(path):
    # pool = threadpool.ThreadPool(20)
    # openpyxl_write(path=path, star_index=0, end_index=499, pool=pool)

    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.worksheets[0]
    # 获取行数
    rows = worksheet.max_row
    worksheet = None
    workbook = None
    # 次数
    i = rows / 500
    print  i
    pool = threadpool.ThreadPool(20)
    for index in range(3, i + 1):
        openpyxl_write(path=path, star_index=(index - 1) * 500, end_index=index * 500-1, pool=pool)
    if rows - i * 500 > 0:
        openpyxl_write(path=path, star_index=i * 500, end_index=rows, pool=pool)


if __name__ == '__main__':
    reload(sys)
    sys.setdefaultencoding('utf-8')
    bootstarp("./data0.xlsx");
    # openpyxl_write("./data0.xlsx")
